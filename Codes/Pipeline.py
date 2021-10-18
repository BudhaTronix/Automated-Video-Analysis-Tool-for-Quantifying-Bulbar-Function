import os
import datetime

from os import path
from openpyxl import load_workbook
from Codes.src.Stabilization import Stabilization
from Codes.src.VideoCompression import VideoCompression
from Codes.src.FaceExtraction import FaceExtraction
from Codes.src.LipExtraction import LipExtraction
from Codes.src.FrequencyCalculator import FrequencyCalculation


def excelUpadter(timestamp, filename, time_slice, frequency_total, errors, sweeps_mode, time_mode, sweeps_mean
                 , time_mean, stdev):
    ExcelFile = "Frequency.xlsx"

    workbook = load_workbook(ExcelFile)
    # workbook.sheetnames
    sheet = workbook.active

    max_column = sheet.max_column
    max_row = sheet.max_row
    sheet.cell(row=max_row + 1, column=1).value = timestamp
    sheet.cell(row=max_row + 1, column=2).value = filename.split("_")[0]
    sheet.cell(row=max_row + 1, column=3).value = time_slice
    sheet.cell(row=max_row + 1, column=4).value = frequency_total
    sheet.cell(row=max_row + 1, column=5).value = errors
    sheet.cell(row=max_row + 1, column=6).value = sweeps_mode
    sheet.cell(row=max_row + 1, column=7).value = sweeps_mean
    sheet.cell(row=max_row + 1, column=8).value = time_mode
    sheet.cell(row=max_row + 1, column=9).value = time_mean
    sheet.cell(row=max_row + 1, column=10).value = stdev

    workbook.save(ExcelFile)


def callStabilize(MainFile):
    print("Starting Video Stabilization..\n")
    Filename = ""
    New_FileName = MainFile.split(".")[0] + "_Stabilized.mp4"
    if path.exists(New_FileName):
        print("Removing existing file...")
        os.remove(New_FileName)
    if path.exists(MainFile):
        Filename = MainFile
    else:
        print("File Not Found")
    Stabilization(Filename, New_FileName)
    print("Video Stabilization Done..\n")


def callCompress(MainFile):
    print("\nStarting Video Compression..\n")
    New_FileName = MainFile.split(".")[0] + "_Compressed.mp4"
    if path.exists(New_FileName):
        print("Removing existing file...")
        os.remove(New_FileName)
    if path.exists(MainFile.split(".")[0] + "_Stabilized.mp4"):
        Filename = MainFile.split(".")[0] + "_Stabilized.mp4"
    else:
        Filename = MainFile
    compress = VideoCompression(Filename, New_FileName)
    if compress:
        print("Video Compression Done..\n")


def callFaceDetect(MainFile, correctionFactor_Face):
    print("\nStarting Face Extraction...\n")
    New_FileName = MainFile.split(".")[0] + "_FaceDetector.mp4"
    if path.exists(New_FileName):
        print("Removing existing file...")
        os.remove(New_FileName)
    if path.exists(MainFile.split(".")[0] + "_Compressed.mp4"):
        Filename = MainFile.split(".")[0] + "_Compressed.mp4"
    elif path.exists(MainFile.split(".")[0] + "_Stabilized.mp4"):
        Filename = MainFile.split(".")[0] + "_Stabilized.mp4"
    else:
        Filename = MainFile
    FaceExtraction(Filename, New_FileName, correctionFactor_Face)
    print("Face Extraction Done...\n")


def callLipExtraction(MainFile, correctionFactor_Lip):
    print("\nStarting Lip Extraction..\n")
    New_FileName = MainFile.split(".")[0] + "_LipDetector.mp4"
    if path.exists(New_FileName):
        print("Removing existing file...")
        os.remove(New_FileName)
    if path.exists(MainFile.split(".")[0] + "_FaceDetector.mp4"):
        Filename = MainFile.split(".")[0] + "_FaceDetector.mp4"
    elif path.exists(MainFile.split(".")[0] + "_Compressed.mp4"):
        Filename = MainFile.split(".")[0] + "_Compressed.mp4"
    elif path.exists(MainFile.split(".")[0] + "_Stabilized.mp4"):
        Filename = MainFile.split(".")[0] + "_Stabilized.mp4"
    else:
        Filename = MainFile
    LipExtraction(Filename, New_FileName, correctionFactor_Lip)
    print("Lip Extraction Done..\n")


def callTongueTrack(MainFile, threshold, thresh_iterations, disp, visual_area, time_slice, model, save_in_excel):
    print("###############Frequency Calculation####################\n")
    Filename = ""
    if path.exists(MainFile.split(".")[0] + "_LipDetector.mp4"):
        Filename = MainFile.split(".")[0] + "_LipDetector.mp4"
    elif path.exists(MainFile.split(".")[0] + "_FaceDetector.mp4"):
        Filename = MainFile.split(".")[0] + "_FaceDetector.mp4"
    else:
        print("File does not exist")
    frequency_total, errors, sweeps_mode, sweeps_mean, stdev, fps, time_slice = FrequencyCalculation(Filename
                                                                                                     , threshold,
                                                                                                     thresh_iterations,
                                                                                                     visual_area,
                                                                                                     disp,
                                                                                                     time_slice,
                                                                                                     model)

    if save_in_excel:
        timestamp = datetime.datetime.now()
        time_mode = (sweeps_mode / fps) * 100
        time_mean = (sweeps_mean / fps) * 100
        excelUpadter(timestamp, Filename, time_slice, frequency_total, errors, sweeps_mode, time_mode, sweeps_mean
                     , time_mean, stdev)
    print("#############Frequency Calcualtion Ends###############\n")


def callConfigEditor():
    print("Starting to Calculate Speed..\n")
    # loaderFunc()
    print("Speed Calculation Done..\n\n")


def callPerformAll():
    print("Starting....\n")
    CUI(MainFile, True, True, True, True, True, SMOOTHING_RADIUS, threshold, thresh_iterations, visual_area, disp
        , correctionFactor_Face, correctionFactor_Lip, model)
    # os.system('python Controller')
    print("Done!\n\n")


def CUI(MainFile, Stabilize, Video_Compression,
        Face_Extract, Lip_Extraction, Frquency_Calculation, threshold, thresh_iterations, visual_area,
        disp, correctionFactor_Face, correctionFactor_Lip, save_in_excel, time_slice, model):
    # Call Stabilization
    print(
        "Stabilization = {},Video_Compression = {},Face_Extraction = {},Lip_Extraction = {},Frequency_Calculation = {}".format
        (Stabilize, Video_Compression, Face_Extract, Lip_Extraction, Frquency_Calculation))
    print("\n")
    if Stabilize:
        callStabilize(MainFile)

    # Call Video Compression
    if Video_Compression:
        callCompress(MainFile)

    # Call Face Extraction
    if Face_Extract:
        callFaceDetect(MainFile, correctionFactor_Face)

    # Call Lip Extraction
    if Lip_Extraction:
        callLipExtraction(MainFile, correctionFactor_Lip)

    # Call Frequency Calculation
    if Frquency_Calculation:
        callTongueTrack(MainFile, threshold, thresh_iterations, disp, visual_area, time_slice, model, save_in_excel)
